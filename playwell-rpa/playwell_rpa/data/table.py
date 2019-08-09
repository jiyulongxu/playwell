"""Table是Playwell RPA中的基本类型
通常应用于不同工作单元之间数据格式的适配，比如要将爬虫单元抓取的数据自动保存为Excel，
那么爬虫单元所输出的数据就是以Table形式存在的，而Excel单元输入的数据也是以Table形式存在的，
两者就可以进行衔接。Table定义了一组抽象协议，在此协议之下可以拥有基于多种存储的实现，比如基于文件、数据库等等。
"""
import time
import os.path
from abc import (
    ABCMeta,
    abstractmethod
)
from typing import Sequence
from openpyxl import (
    load_workbook,
    Workbook
)

from playwell.service import (
    Result,
    PlaywellServiceException,
    single_request_service
)
from playwell.service.message import (
    Message,
    ServiceRequestMessage
)
from playwell.service.message.bus import message_bus_manager
from playwell_rpa.resource import resource_tracer


class Column:

    """Table column元信息
    """

    def __init__(
        self,
        name: str,
        title: str = None,
        primary_key: bool = False,
        type: str = None,
        auto_trans_type: bool = False,
        allow_null: bool = True,
        unique: bool = False,
        max_length: int = None,
        auto_increment: tuple = None
    ):
        """
        :param name: column名称
        :param title:  column title 用于展示
        :param primary_key: column是否为主键
        :param type: 数据类型
        :param auto_trans_type: 是否自动转化数据类型
        :param allow_null:  是否允许空值
        :param unique:  是否唯一
        :param max_length: 最大长度
        :param auto_increment: 自动递增
        """
        self._name = name
        if title is None:
            title = name
        self._title = title
        self._primary_key = primary_key
        self._type = type
        self._auto_trans_type = auto_trans_type
        self._allow_null = allow_null
        self._unique = unique
        self._max_length = max_length
        self._auto_increment = auto_increment

    @property
    def name(self):
        return self._name

    @property
    def title(self):
        return self._title

    @property
    def primary_key(self):
        return self._primary_key

    @property
    def type(self):
        return self._type

    @property
    def auto_trans_type(self):
        return self._auto_trans_type

    @property
    def allow_null(self):
        return self._allow_null

    @property
    def unique(self):
        return self._unique

    @property
    def max_length(self):
        return self._max_length

    @property
    def auto_increment(self):
        return self._auto_increment


class Row:

    """Table Row
    """

    def __init__(self, elements: Sequence, col_names: Sequence[str], mapping: dict):
        elements = elements[:len(col_names)]
        self._col_names = col_names
        self._mapping = mapping
        self._elements = elements

    def view(self, col_names: Sequence[str]):
        if not col_names:
            return self
        elements = [self[col_name] for col_name in col_names]
        mapping = {col_name: idx for idx, col_name in enumerate(col_names)}
        return Row(elements, col_names, mapping)

    @property
    def elements(self):
        return self._elements

    @property
    def map(self):
        return {self._col_names[idx]: ele for idx, ele in enumerate(self.elements)}

    def __getitem__(self, key: int or str):
        if isinstance(key, int):
            return self.get_by_index(key)
        elif isinstance(key, str):
            index = self._mapping[key]
            return self.get_by_index(index)

    def get_by_index(self, index: int):
        if index >= len(self._elements):
            raise PlaywellServiceException("Invalid row index: %d" % index)
        return self._elements[index]

    def __len__(self):
        return len(self._elements)

    def __iter__(self):
        yield from self._elements

    def __repr__(self):
        return repr(self._elements)

    def __str__(self):
        return str(self._elements)


class Table(metaclass=ABCMeta):

    """Abstract table api
    """

    def __init__(self, name: str, columns: Sequence[Column]):
        """
        :param name: table name
        :param columns: table columns
        """
        self._name = name
        self._columns = list(columns)
        self._col_names = [col.name for col in columns]
        self._mapping = {col.name: idx for idx, col in enumerate(columns)}

    @property
    def name(self):
        return self._name

    @property
    def columns(self):
        return self._columns

    @abstractmethod
    def append_rows(self, rows):
        pass

    @abstractmethod
    def clear(self):
        pass

    @abstractmethod
    def __len__(self):
        pass

    @abstractmethod
    def __iter__(self):
        pass


class MemoryTable(Table):

    ALL_MEM_TABLES = {}

    @classmethod
    def register_table(
        cls,
        activity_id: int,
        domain_id: str,
        name: str,
        columns: Sequence[Column],
        _meta: dict
    ):
        full_name = _full_table_name(activity_id, domain_id, name)
        table = cls(full_name, columns)
        MemoryTable.ALL_MEM_TABLES[full_name] = table
        return table

    @staticmethod
    def locate_table(
        activity_id: int,
        domain_id: str,
        name: str,
        _meta: dict
    ):
        full_name = _full_table_name(activity_id, domain_id, name)
        if full_name not in MemoryTable.ALL_MEM_TABLES:
            raise PlaywellServiceException("找不到指定的Memory table: %s" % name)
        return MemoryTable.ALL_MEM_TABLES[full_name]

    @staticmethod
    def drop_table(
        activity_id: int,
        domain_id: str,
        name: str
    ):
        full_name = _full_table_name(activity_id, domain_id, name)
        if full_name not in MemoryTable.ALL_MEM_TABLES:
            raise PlaywellServiceException("找不到指定的Memory table: %s" % name)
        MemoryTable.ALL_MEM_TABLES[full_name].clear()
        del MemoryTable.ALL_MEM_TABLES[full_name]

    def __init__(self, name: str, columns: Sequence[Column]):
        super().__init__(name, columns)
        self._data = []

    def append_rows(self, rows):
        self._data.extend([Row(row_data, self._col_names, self._mapping) for row_data in rows])

    def clear(self):
        self._data = []

    def __len__(self):
        return len(self._data)

    def __iter__(self):
        yield from self._data


class ExcelTable(Table):

    ALL_EXCEL_TABLE_META = {}

    @classmethod
    def register_table(
        cls,
        activity_id: int,
        domain_id: str,
        name: str,
        columns: Sequence[Column],
        meta: dict
    ):
        full_name = _full_table_name(activity_id, domain_id, name)
        file_path, sheet_name, has_title = meta["path"], meta["sheet"], meta["has_title"]

        if os.path.exists(file_path):
            work_book = load_workbook(filename=file_path)
        else:
            work_book = Workbook()
            ws1 = work_book.active  # 移除新建Excel的默认sheet
            work_book.remove(ws1)

        try:
            if sheet_name in work_book:
                sheet = work_book[sheet_name]
                if not columns:
                    if not has_title:
                        raise PlaywellServiceException(
                            "Excel table %s 没有定义columns，并且也无法从表格自动获取" % name)
                    columns = [Column(name=cell.value) for cell in sheet[1] if cell.value]
            else:
                work_sheet = work_book.create_sheet(title=sheet_name)
                if has_title:
                    work_sheet.append([col.title for col in columns])  # 向新的sheet中写入标题
                work_book.save(file_path)
        finally:
            work_book.close()

        table = cls(
            name=name,
            columns=columns,
            file_path=file_path,
            sheet_name=sheet_name,
            has_title=has_title
        )
        ExcelTable.ALL_EXCEL_TABLE_META[full_name] = table
        return table

    @staticmethod
    def locate_table(
        activity_id: int,
        domain_id: str,
        name: str,
        _meta: dict
    ):
        full_name = _full_table_name(activity_id, domain_id, name)
        if full_name not in ExcelTable.ALL_EXCEL_TABLE_META:
            raise PlaywellServiceException("找不到指定的Excel table: %s" % name)
        return ExcelTable.ALL_EXCEL_TABLE_META[full_name]

    @staticmethod
    def drop_table(
        activity_id: int,
        domain_id: str,
        name: str
    ):
        full_name = _full_table_name(activity_id, domain_id, name)
        if full_name not in ExcelTable.ALL_EXCEL_TABLE_META:
            raise PlaywellServiceException("找不到指定的Excel table: %s" % name)
        del ExcelTable.ALL_EXCEL_TABLE_META[full_name]

    def __init__(
        self,
        name: str,
        columns: Sequence[Column],
        file_path: str,
        sheet_name: str,
        has_title: bool
    ):
        super().__init__(name, columns)
        self._file_path = file_path
        self._sheet_name = sheet_name
        self._has_title = has_title

    @property
    def file_path(self):
        return self._file_path

    @property
    def sheet_name(self):
        return self._sheet_name

    def append_rows(self, rows):
        work_book, work_sheet = self._get_work_book_and_sheet()
        try:
            for row in rows:
                work_sheet.append(row)
            work_book.save(self._file_path)
        finally:
            work_book.close()

    def clear(self):
        raise PlaywellServiceException("Excel table not support this operation!")

    def __len__(self):
        work_book, work_sheet = self._get_work_book_and_sheet()
        try:
            if self._has_title:
                return work_sheet.max_row - 1
            else:
                return work_sheet.max_row
        finally:
            work_book.close()

    def __iter__(self):
        work_book, work_sheet = self._get_work_book_and_sheet()
        try:
            for idx, row_data in enumerate(work_sheet):
                if self._has_title and idx == 0:
                    continue

                elements = [cell.value for cell in row_data]
                if any(elements):  # 非空行
                    yield Row(elements, self._col_names, self._mapping)
                else:
                    continue
        finally:
            work_book.close()

    def _get_work_book_and_sheet(self):
        if not os.path.exists(self._file_path):
            raise PlaywellServiceException(
                "Could not found the excel file: %s" % self._file_path)
        work_book = load_workbook(filename=self._file_path)
        if self._sheet_name not in work_book:
            raise PlaywellServiceException(
                "找不到sheet %s 在workbook: %s" % (self._sheet_name, self._file_path))
        return work_book, work_book[self._sheet_name]


_table_types = {
    "memory_table": MemoryTable,
    "excel_table": ExcelTable
}

for (type, class_type) in _table_types.items():
    resource_tracer.register_handler(type, getattr(class_type, "drop_table"))


def declare_table(
    type: str,
    activity_id: int,
    domain_id: str,
    name: str,
    columns: Sequence[Column],
    meta: dict,
    tmp: bool
):
    global _table_types
    _check_table_type(type)
    text = _table_types[type].register_table(activity_id, domain_id, name, columns, meta)
    if tmp:
        resource_tracer.trace(
            activity_id=activity_id,
            domain_id=domain_id,
            type=type,
            name=name
        )
    return text


def locate_table(type: str, activity_id: int, domain_id: str, name: str, meta: dict):
    global _table_types
    _check_table_type(type)
    return _table_types[type].locate_table(activity_id, domain_id, name, meta)


def locate_from_request(ref_name: str, request: ServiceRequestMessage):
    args = request.args
    table_type, table_name, meta, _tmp = args[ref_name]
    return locate_table(
        type=table_type,
        activity_id=request.activity_id,
        domain_id=request.domain_id,
        name=table_name,
        meta=meta
    )


def drop_table(activity_id: int, domain_id: str, type: str, name: str):
    global _table_types
    _check_table_type(type)
    return _table_types(type).drop_table(activity_id, domain_id, name)


def _check_table_type(type: str):
    if type not in _table_types:
        raise PlaywellServiceException("未知的Table类型: %s" % type)


def _full_table_name(activity_id: int, domain_id: str, name: str):
    return "%d:%s:%s" % (activity_id, domain_id, name)


@single_request_service
def declare_new_table(request: ServiceRequestMessage):
    """声明表格的元信息
    - name: create_table
      type: table.declare
      args:
        request:
          table: ref("excel_table", "students")
          columns:
            - name: str("id")
              title: str("ID")
            - name: str("name")
              title: str("姓名")
          meta:
            path: str("/tmp/test.xlsx")
            sheet: str("sheet")
            has_title: yes
    """
    args = request.args
    type, name, _, tmp = args["table"]
    columns = [Column(**col_data) for col_data in args.get("columns", [])]
    meta = args.get("meta", {})
    declare_table(
        type=type,
        activity_id=request.activity_id,
        domain_id=request.domain_id,
        name=name,
        columns=columns,
        meta=meta,
        tmp=tmp
    )
    return Result.ok()


@single_request_service
def table_to_list(request: ServiceRequestMessage):
    """将表格转化成list，并在Result中返回
    - name: table.to_list
      args:
        request:
          table: ref("excel_table", "students")
          row_style: str("list")
          columns:
            - str("id")
            - str("name")
        timeout: minutes(1)
      ctrl:
        - when: resultOk()
          context_vars:
            list: resultVar("list")
          then: call("next_action")
        - default: failBecause("to_list_failure")
    """
    table = locate_from_request("table", request)
    row_style = request.args.get("row_style", "map")
    col_names = request.args.get("columns", [])

    if table:
        lst = []
        for row in table:
            row = row.view(col_names)
            if row_style == "list":
                lst.append(row.elements)
            elif row_style == "map":
                lst.append(row.map)
            else:
                return Result.fail(
                    error_code="invalid_row_style",
                    message=("Invalid row style: %s" % row_style)
                )
    else:
        lst = []

    return Result.ok(data={
        "list": lst
    })


@single_request_service
def table_to_messages(request: ServiceRequestMessage):
    """可将表格转化为消息的服务
    - name: table.to_messages
      args:
        request:
          table: ref("memory_table", "students")
          message_bus: str("target_message_bus")
          message_type: str("notify_student")
          columns:
            - str("ID")
            - str("Name")
            - str("Phone")
          batch_size: 1000
          sleep: 5
    """
    args = request.args
    table = locate_from_request("table", request)
    message_type, message_bus_name = args["message_type"], args["message_bus"]
    col_names = args.get("columns", [])
    batch_size, sleep = args.get("batch_size", 1000), args.get("sleep", 0)

    message_bus = message_bus_manager.get_message_bus(message_bus_name)
    if message_bus is None:
        return Result.fail(
            error_code="bus_not_found",
            message="Could not found message bus: %s" % message_bus_name
        )

    messages = []
    ts = int(time.time() * 1000)
    for idx, row in enumerate(table):
        messages.append(Message(
            type=message_type,
            sender="",
            receiver="",
            attr=row.view(col_names).map,
            time=ts
        ))

        if idx % batch_size == 0:
            message_bus.write(messages)
            messages = []
            ts = int(time.time() * 1000)

            if sleep != 0:
                time.sleep(sleep)

    if messages:
        message_bus.write(messages)

    return Result.ok()
